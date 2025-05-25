#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Dec 30 13:30:30 2024

@author: Jason G. Karlin
karlin@iii.u-tokyo.ac.jp
"""
### GirlsinSight
### version 1.0.1

from openai import OpenAI
import anthropic
import google.generativeai as genai  # Add this import
import pandas as pd
import requests
from bs4 import BeautifulSoup
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import time
import sys
from urllib.parse import urljoin
import re
import datetime
from datetime import timedelta
import locale
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import matplotlib.dates as mdates
from matplotlib.colors import LinearSegmentedColormap
import seaborn as sns
from urllib.parse import quote_plus
import warnings
import logging
import os
from google.colab import userdata
import subprocess
import asyncio
import aiohttp
import nest_asyncio

os.environ["LANG"] = "ja_JP.UTF-8"

# Suppress only specific warnings while keeping important messages
warnings.simplefilter(action='ignore', category=FutureWarning)
os.environ['IMK_DISABLE_LOGGING'] = '1'
os.environ['GRPC_ENABLE_FORK_SUPPORT'] = '0'

# Suppress Google API client messages
logging.getLogger('googleapiclient.discovery_cache').setLevel(logging.ERROR)
logging.getLogger('googleapiclient.discovery').setLevel(logging.ERROR)

# Get the current date and time
current_datetime = datetime.datetime.now()
# Format the date string - this will work even if locale setting fails
formatted_datetime = f"{current_datetime.year}年{current_datetime.month}月{current_datetime.day}日 {current_datetime.hour}:{current_datetime.minute:02d}:{current_datetime.second:02d}"

# Create the "outputs" subfolder if it doesn't exist
output_folder = "outputs"
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

# Get API keys from Colab secrets
try:
    client_gpt = OpenAI(
        api_key=userdata.get('OPENAI_API_KEY')
    )
except:
    client_gpt = None

try:
    client_anthropic = anthropic.Client(
        api_key=userdata.get('ANTHROPIC_API_KEY')
    )
except:
    client_anthropic = None

# Required API key for Gemini
genai.configure(api_key=userdata.get('GEMINI_API_KEY'))
client_gemini = genai.GenerativeModel('gemini-2.5-pro-preview-05-06')

# Get the directory of the current script
directory_path = os.path.dirname(os.path.realpath(__file__))

# Ensure the script runs in Colab's default directory
os.chdir('/content')

# Define headers for web requests
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.150 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
    'Accept-Language': 'ja,en-US;q=0.9,en;q=0.8',
    'Accept-Encoding': 'gzip, deflate, br',
    'Connection': 'keep-alive',
    'Referer': 'https://www.google.com/'
}

class TopicCounter:
    plot_results_count = 1
    plot_comment_frequency_count = 1

def setup_japanese_fonts():
    import matplotlib.pyplot as plt
    import matplotlib.font_manager as fm
    import logging
    import subprocess
    
    # Suppress matplotlib font manager warnings
    logging.getLogger('matplotlib.font_manager').setLevel(logging.ERROR)
    
    # Install Japanese fonts using subprocess
    try:
        subprocess.run(['apt-get', 'install', '-y', 'fonts-noto-cjk', 'fonts-noto-cjk-extra'], 
                      check=True, 
                      capture_output=True)
    except subprocess.CalledProcessError:
        print("Warning: Unable to install Japanese fonts. Text display may be affected.")
    
    # Specify the path to the Japanese font
    font_path = "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"
    font_prop = fm.FontProperties(fname=font_path)
    
    # Set font globally
    plt.rcParams['font.family'] = font_prop.get_name()
    
    return font_prop

def create_figure_with_japanese_fonts():
    import matplotlib.pyplot as plt
    import matplotlib.font_manager as fm
    import logging
    
    # Suppress matplotlib font manager warnings
    logging.getLogger('matplotlib.font_manager').setLevel(logging.ERROR)
    
    # Specify the path to the Japanese font
    font_path = "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"
    font_prop = fm.FontProperties(fname=font_path)
    
    # Set font globally
    plt.rcParams['font.family'] = font_prop.get_name()
    
    # Additional settings
    plt.rcParams['axes.unicode_minus'] = False
    
    return plt.figure(), font_prop

# Function to get search query and date selection from user
def get_user_input():
    # Prompt user for menu choice, URL or search query, and language selection
    print("\n*** GirlsinSight ***")
    while True:
        print("\n1. トピックのURLを入力")
        print("2. トピックを検索")
        print("3. 終了する")
        menu_choice = input("\n選択肢を入力してください: ")
        
        url, comment_total = None, None  # Initialize variables
        
        if menu_choice == '1':
            for attempts in range(3):  # Use for loop for controlled attempts
                url = input("トピックのURLを入力してください: ")
                comment_total = get_comment_total(url)  # This function validates the URL
                if comment_total is not None:
                    while True:
                        language_selection = input("\n結果はどちらの言語がよろしいですか： 1. 日本語 2. 英語： ")
                        if language_selection in ['1', '2']:
                            break
                        else:
                            print("無効な選択です。1または2を入力してください。")
                    
                    while True:
                        summary_ai = input("\nAIを選択してください (要約)： 1. OpenAI 2. Anthropic 3. Gemini： ")
                        if summary_ai in ['1', '2', '3']:
                            break
                        else:
                            print("無効な選択です。1、2、または3を入力してください。")
                    
                    while True:
                        sentiment_ai = input("\nAIを選択してください (感情分析)： 1. OpenAI 2. Anthropic 3. Gemini： ")
                        if sentiment_ai in ['1', '2', '3']:
                            break
                        else:
                            print("無効な選択です。1、2、または3を入力してください。")
                    
                    # Assign respective values based on user input for summary AI
                    if summary_ai == '1':
                        summary_ai = 'gpt'
                    elif summary_ai == '2':
                        summary_ai = 'anthropic'
                    elif summary_ai == '3':
                        summary_ai = 'gemini'
                    
                    # Assign respective values based on user input for sentiment AI
                    if sentiment_ai == '1':
                        sentiment_ai = 'gpt'
                    elif sentiment_ai == '2':
                        sentiment_ai = 'anthropic'
                    elif sentiment_ai == '3':
                        sentiment_ai = 'gemini'
                    
                    return menu_choice, url, language_selection, comment_total, summary_ai, sentiment_ai
                else:
                    print("有効なURLが入力されませんでした。")
            return None, None, None, None, None, None  # Return None values to indicate an invalid URL
        
        elif menu_choice == '2':
            search_query = input("検索キーワードを入力ください: ")
            while True:
                print("\n日付範囲:")
                print("1. 全期間")
                print("2. １年以内")
                print("3. １ヶ月以内")
                print("4. １週間以内")
                date_selection = input("\n選択した番号を入力してください: ")
                search_url = construct_search_url(search_query, date_selection)
                result = fetch_topic_urls(search_url)
                if result:
                    topic_urls, comment_totals = result
                    num_topics_found = len(topic_urls)
                    if topic_urls:
                        print(f"\nトピックの総数: {num_topics_found}\n")
                        while True:
                            proceed = input("続行しますか？ (y/n): ")
                            if proceed.lower() == 'y':
                                while True:
                                    language_selection = input("\n結果はどちらの言語がよろしいですか： 1. 日本語 2. 英語： ")
                                    if language_selection in ['1', '2']:
                                        break
                                    else:
                                        print("無効な選択です。1または2を入力してください。")
                                
                                while True:
                                    summary_ai = input("\nAIを選択してください (要約)： 1. OpenAI 2. Anthropic 3. Gemini： ")
                                    if summary_ai in ['1', '2', '3']:
                                        break
                                    else:
                                        print("無効な選択です。1、2、または3を入力してください。")
                                
                                while True:
                                    sentiment_ai = input("\nAIを選択してください (感情分析)： 1. OpenAI 2. Anthropic 3. Gemini： ")
                                    if sentiment_ai in ['1', '2', '3']:
                                        break
                                    else:
                                        print("無効な選択です。1、2、または3を入力してください。")
                                
                                # Assign respective values based on user input for summary AI
                                if summary_ai == '1':
                                    summary_ai = 'gpt'
                                elif summary_ai == '2':
                                    summary_ai = 'anthropic'
                                elif summary_ai == '3':
                                    summary_ai = 'gemini'
                                
                                # Assign respective values based on user input for sentiment AI
                                if sentiment_ai == '1':
                                    sentiment_ai = 'gpt'
                                elif sentiment_ai == '2':
                                    sentiment_ai = 'anthropic'
                                elif sentiment_ai == '3':
                                    sentiment_ai = 'gemini'
                                
                                return menu_choice, search_query, topic_urls, comment_totals, num_topics_found, language_selection, summary_ai, sentiment_ai
                            elif proceed.lower() == 'n':
                                break  # Break out of the inner loop and continue with the outer loop
                            else:
                                print("無効な入力です。'y' または 'n' を入力してください。")
                    else:
                        print(f"「{search_query}」に該当するトピックはありませんでした。")
                        break  # Break out of the loop and return to the start
                else:
                    if date_selection == '1':
                        print(f"「{search_query}」に該当するトピックはありませんでした。")
                        break  # Break out of the loop and return to the start
        elif menu_choice == '3':
            print("メニューから退出します。")
            sys.exit()
        else:
            print("無効な選択肢です。1, 2, または 3 を入力してください。")
            sys.exit()

def process_dates(group_df):
    earliest_date = None
    latest_date = None
    
    for date in group_df["Date"]:
        try:
            if earliest_date is None or date < earliest_date:
                earliest_date = date
            if latest_date is None or date > latest_date:
                latest_date = date
        except (ValueError, TypeError):
            continue  # Skip rows where the date cannot be parsed
    
    # Calculate the duration
    delta = latest_date - earliest_date
    days = delta.days
    seconds = delta.seconds
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    seconds = seconds % 60
    
    return {'earliest': earliest_date, 'latest': latest_date, 'days': days, 'hours': hours, 'minutes': minutes, 'seconds': seconds}

def search_wayback_machine(url, headers, max_retries=3, retry_delay=1):
    """
    Searches the Wayback Machine for a working archived version of a URL.
    Prioritizes finding the oldest valid snapshot from recent CDX results,
    then falls back to the 'closest' available snapshot.
    
    Args:
        url (str): The URL to search for
        headers (dict): Request headers
        max_retries (int): Maximum number of retry attempts for failed requests
        retry_delay (int): Delay in seconds between retries
    
    Returns:
        str or None: Returns the working Wayback Machine URL if found, None otherwise
    """
    def make_request(search_url, attempt=0):
        try:
            response = requests.get(search_url, headers=headers, timeout=10)
            response.raise_for_status()
            return response.json()
        except (requests.RequestException, ValueError) as e:
            # Reduced verbosity for retries unless max_retries is hit
            if attempt < max_retries -1 : # try silently until the last attempt
                 time.sleep(retry_delay)
                 return make_request(search_url, attempt + 1)
            # On last attempt, print error if it fails
            if attempt == max_retries -1:
                time.sleep(retry_delay)
                try:
                    response = requests.get(search_url, headers=headers, timeout=10)
                    response.raise_for_status()
                    return response.json()
                except (requests.RequestException, ValueError) as final_e:
                    print(f"Error fetching {search_url} after {max_retries} attempts: {str(final_e)}")
            return None

    def verify_url(wayback_url_to_check):
        try:
            # Use HEAD request, allow redirects, and check Content-Type
            verify_response = requests.head(
                wayback_url_to_check,
                headers=headers,
                timeout=15,
                allow_redirects=True  # Important: follow redirects
            )
            verify_response.raise_for_status()  # Check for 2xx status codes after redirects

            content_type = verify_response.headers.get('Content-Type', '').lower()
            is_html = 'text/html' in content_type
            
            return is_html
        except requests.RequestException:
            # Catches timeouts, connection errors, too_many_redirects, non-2xx from raise_for_status
            return False

    # Attempt 1: CDX API for a list of snapshots, aiming for the oldest valid one among recent captures
    # Get up to 20 recent captures that were originally 200 OK and text/html
    cdx_api_url = f"http://web.archive.org/cdx/search/cdx?url={quote_plus(url)}&output=json&filter=statuscode:200&filter=mimetype:text/html&limit=20"
    data_cdx = make_request(cdx_api_url)
    
    valid_cdx_snapshots = []
    if data_cdx:
        if len(data_cdx) > 0 and isinstance(data_cdx[0], list) and len(data_cdx[0]) > 1 and data_cdx[0][0] == 'urlkey':
             # Skip header row if present (Wayback CDX API sometimes includes it)
             data_cdx_iterable = data_cdx[1:]
        else:
             data_cdx_iterable = data_cdx

        for snapshot_details in data_cdx_iterable:
            try:
                # Standard CDX format: urlkey, timestamp, original, mimetype, statuscode, digest, length
                # Ensure snapshot_details is a list and has enough elements
                if not isinstance(snapshot_details, list) or len(snapshot_details) < 3:
                    continue

                timestamp = snapshot_details[1]
                original_page_url = snapshot_details[2]
                
                wayback_candidate_url = f"http://web.archive.org/web/{timestamp}/{original_page_url}"
                
                if verify_url(wayback_candidate_url):
                    valid_cdx_snapshots.append({
                        'url': wayback_candidate_url,
                        'timestamp': timestamp 
                    })
            except (IndexError, KeyError, TypeError):
                # Silently skip malformed entries
                continue
    
    if valid_cdx_snapshots:
        valid_cdx_snapshots.sort(key=lambda s: s['timestamp']) # Sort oldest first
        # print(f"Debug: Found {len(valid_cdx_snapshots)} valid snapshots from CDX. Oldest: {valid_cdx_snapshots[0]['url']}")
        return valid_cdx_snapshots[0]['url']

    # Attempt 2: Fallback to standard "available" API for the 'closest' snapshot
    available_url_api = f"http://archive.org/wayback/available?url={quote_plus(url)}"
    data_available = make_request(available_url_api)
    
    if data_available and 'archived_snapshots' in data_available and \
       'closest' in data_available['archived_snapshots'] and \
       data_available['archived_snapshots']['closest'] and \
       'url' in data_available['archived_snapshots']['closest']:
        
        closest_url = data_available['archived_snapshots']['closest']['url']
        if closest_url and verify_url(closest_url):
            # print(f"Debug: Falling back to 'closest' URL from available API: {closest_url}")
            return closest_url

    # print(f"Debug: No working archive found for {url}")
    return None

# Get the comment count for single url topic
def get_comment_total(url):
    try:
        response = requests.get(url, headers = headers)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'html.parser')
            comment_span = soup.find('span', class_='icon-comment')
            if comment_span:
                next_span = comment_span.find_next_sibling('span')
                if next_span:
                    comment_total = next_span.text.replace('コメント', '')
                    comment_total = int(comment_total)  # Convert to integer
                    return comment_total
                else:
                    print("コメント数が見つかりませんでした。")  # Print if comment count not found
                    return None
            else:
                print("コメント数が見つかりませんでした。")  # Print if comment count not found
                return None
        else:
            print("URLの取得に失敗しました。")  # Print if URL request fails
            return None
    except requests.exceptions.RequestException:
        print("入力されたURLは無効です。正しいURLをご確認の上、もう一度お試しください。\n")
        return None

def get_page_source(gc_url, chunk_size=102400):
    response = requests.get(gc_url)
    response.encoding = 'utf-8'
    page_source_all = BeautifulSoup(response.content, 'html.parser')

    # Take only the first chunk of the specified size from the BeautifulSoup object
    page_source = BeautifulSoup(str(page_source_all)[:chunk_size], 'html.parser')

    return page_source

def create_url_list(url_to_process):
    base_url = "https://girlschannel.net"  # Or extract from the original URL
    all_urls = [url_to_process]
    gc_url = url_to_process
    while gc_url:
        try:
            page_source = get_page_source(gc_url)
            next_page_url = get_nextpage_link(page_source)
            
            if next_page_url != "null":
                # Convert relative URL to absolute URL
                if next_page_url.startswith('/'):
                    # Check if this is a domain-relative URL (starts with /)
                    # Extract the domain from the original URL
                    from urllib.parse import urlparse
                    parsed_original = urlparse(url_to_process)
                    base_domain = f"{parsed_original.scheme}://{parsed_original.netloc}"
                    next_page_url = base_domain + next_page_url
                elif not next_page_url.startswith(('http://', 'https://')):
                    # If it's a page-relative URL (doesn't start with / or http)
                    # Combine with the path of the current URL
                    next_page_url = urljoin(gc_url, next_page_url)
                
                if next_page_url in all_urls:
                    print("No additional pages found.")
                    break
                else:
                    all_urls.append(next_page_url)
                    gc_url = next_page_url
            else:
                print("Single page article detected.")
                break
        except requests.ConnectionError:
            print(f"\nConnectionError encountered. \nTrying to find {gc_url} on the Wayback Machine...")
            wayback_url = search_wayback_machine(gc_url, headers)
            if wayback_url:
                print("\nアーカイブ版が見つかりました。")
                all_urls = [wayback_url]  # Replace all_urls with the Wayback Machine URL
                break
            else:
                print("\nウェイバックマシンにアーカイブ版が見つかりませんでした。")
                break
        except requests.exceptions.RequestException as e:
            print(f"\nError occurred while processing {gc_url}: {str(e)}")
            print(f"\nTrying to find {gc_url} on the Wayback Machine...")
            wayback_url = search_wayback_machine(gc_url, headers)
            if wayback_url:
                print("\nアーカイブ版が見つかりました。")
                all_urls = [wayback_url]  # Replace all_urls with the Wayback Machine URL
                break
            else:
                print("\nウェイバックマシンにアーカイブ版が見つかりませんでした。")
                break
    
    print("\nAll URLs scraped:", all_urls)
    if len(all_urls) == 1:
        print("Single page article confirmed - proceeding with processing...")
    return all_urls

# Extract text from all URLs in the list
def extract_text_from_url(all_urls, headers):
    article_text = []
    for gc_url in all_urls:
        try:
            response = requests.get(gc_url, headers=headers)
            # Force encoding to UTF-8 or try to detect the correct encoding
            response.encoding = response.apparent_encoding or 'utf-8'
            soup = BeautifulSoup(response.content.decode(response.encoding, errors='replace'), 'html.parser')
            article_text.append(' '.join(p.get_text() for p in soup.find_all('p')))
        except requests.exceptions.RequestException as e:
            print(f"Error extracting text from URL: {gc_url}")
            print(f"Error message: {str(e)}")
    return ' '.join(article_text)

#Analyze page for next page URL
def get_nextpage_link(page_source):
    # Initialize the model
    model = genai.GenerativeModel('gemini-2.5-pro-preview-05-06')
    
    # Create the prompt
    prompt = f"""You are an advanced HTML parser tasked with extracting specific URLs from HTML content. Your primary goal is to find either a 'full text' link or a 'next page' link for an article. Please follow these steps to analyze the HTML and extract the required URL:

1. Search for a 'full text' link:
   - Look for links containing text or attributes related to 'full text' or '記事全文' (Japanese for 'full text of the article').
   - If found, this link takes priority over any pagination links.

2. If no 'full text' link is found, check for pagination:
   - Look for elements indicating that the article is split across multiple pages.
   - Search for classes or IDs containing keywords like 'next', 'next-page', '次', or '続き'.
   - Also, check for navigation arrows or sequential numbers that might indicate pagination.

3. If neither a 'full text' link nor pagination is found, prepare to return 'null'.

4. URL Handling:
   - For any found URL, check if it's a relative URL (starts with '/' or doesn't contain 'http').
   - If it's a relative URL, combine it with the base URL using the following rules:
     a. If the URL starts with '/', append it to the base domain
     b. If the URL is relative to the current path, resolve it against the current page URL
   - Always return a complete, absolute URL

Output Format:
After your analysis, provide **only** one of the following:
1. The URL of the 'full text' link (if found)
2. The URL of the 'next page' link (if pagination is found)
3. The word 'null' (if neither is found)

Here is the HTML content you need to analyze: {page_source}"""

    # Generate response
    response = model.generate_content(
        prompt,
        generation_config=genai.GenerationConfig(
            max_output_tokens=500,
            temperature=0.1,
        )
    )
    
    return response.text.strip()

def check_news_story_status(url_to_process):
    gc_url = url_to_process
    page_source_for_analysis = get_page_source(gc_url)
    
    # Initialize the model
    model = genai.GenerativeModel('gemini-2.5-pro-preview-05-06')
    
    # Create the prompt
    prompt = f"""You are an HTML parser.

Please analyze the following HTML source to determine if the news story has been deleted, removed, or expired. Look for indicators such as error messages, '404' status codes, or specific phrases like 'Page not found', '記事が見つかりません', or 'expired'. If the news story is still available, respond with 'active'. If it has been deleted, removed, or expired, respond with 'inactive'. Respond only with either 'active' or 'inactive'.

HTML Content: {page_source_for_analysis}"""

    # Generate response
    response = model.generate_content(
        prompt,
        generation_config=genai.GenerationConfig(
            max_output_tokens=50,
            temperature=0.1,
        )
    )
    
    return response.text.strip()

# Summarize article using Gemini models
import google.generativeai as genai

def summarize_article_with_gemini(article_text, language_selection):
    language_options = {"1": "Japanese", "2": "English"}
    language = language_options.get(language_selection, "")
    
    # Initialize the model
    model = genai.GenerativeModel('gemini-2.5-pro-preview-05-06')
    
    # Create the prompt
    prompt = f"""You are an expert content creator specializing in {language} summaries.

In {language}, please provide a detailed and engaging summary of the following:

{article_text}

In your summary, highlight notable details, quotations, and/or statistics while aiming for a summary that's easy to read yet informative. If any additional unrelated news items, segments, warnings, alerts, stories, notifications, or advertisements arise, then ignore in your summary in favor of the primary topic: {header}. Do not reference these instructions in your response."""

    # Generate response
    response = model.generate_content(
        prompt,
        generation_config=genai.GenerationConfig(
            max_output_tokens=2048,
            temperature=0.7,
        )
    )
    
    return response.text

# Summarize topic using Gemini models
import google.generativeai as genai

def summarize_topic_with_gemini(topcomment_text, language_selection):
    language_options = {"1": "Japanese", "2": "English"}
    language = language_options.get(language_selection, "")
    
    # Initialize the model
    model = genai.GenerativeModel('gemini-2.5-pro-preview-05-06')
    
    # Create the prompt
    prompt = f"""You are a social media analyst specializing in {language} content analysis.

In {language}, please explain the overall meaning and likely intent of the post, combining insights from its title:
{header}

And the accompanying comment:
{topcomment_text}

Provide a clear, concise analysis that focuses on the overall meaning and likely intent. Avoid restating the title and comment directly and do not use introductory phrases."""

    # Generate response
    response = model.generate_content(
        prompt,
        generation_config=genai.GenerationConfig(
            max_output_tokens=1024,
            temperature=0.7,
        )
    )
    
    return response.text

def evaluate_sentiment_with_gemini(search_query, highest_sentiment_comments, 
                                 lowest_sentiment_comments, language_selection, 
                                 comments_to_analyze):
    language_options = {"1": "Japanese", "2": "English"}
    language = language_options.get(language_selection, "")
    
    # Initialize the model inside the function
    model = genai.GenerativeModel('gemini-2.5-pro-preview-05-06')
    
    # Create the prompt
    system_prompt = f"""You are a helpful Japanese assistant trained to analyze comments from the online forum GirlsChannel.net related to {search_query}. Your task is to write a comprehensive, cohesive essay in {language} that evaluates the sentiments expressed by the users. The essay should integrate the main themes, emotions, and attitudes present in the comments, providing a nuanced perspective on the overall community sentiment. Aim to write a well-structured essay of approximately 500 words, focusing on the predominant opinions while also considering dissenting views. The essay should flow smoothly, without being divided into distinct sections.

In {language}, please write a comprehensive, cohesive essay of approximately 500 words analyzing and summarizing the overall sentiment on {search_query} based on the following dataset aggregated from GirlsChannel.net. The dataset includes the {comments_to_analyze} most upvoted comments {highest_sentiment_comments} alongside the {comments_to_analyze} most downvoted comments {lowest_sentiment_comments}, reflecting the community's overall sentiment and opinion. Since the {lowest_sentiment_comments} have been voted down by users, interpret the prevailing opinion or sentiment to be in strong disagreement with those comments.

Your essay should flow smoothly from one idea to the next, without distinct sections or headers. Evaluate the upvoted and downvoted comments together, integrating the sentiments and themes they express to discern the collective opinion on {search_query}. Focus on the predominant emotions, attitudes, and topics to understand what aspects of {search_query} resonate most with the GirlsChannel.net online community.

Address the main themes or topics that emerge from the comments, the prevailing emotions and attitudes expressed by the community, how the most upvoted and downvoted comments differ in their sentiments, any notable quotes or specific comments that encapsulate the overall sentiment, and how the sociocultural context of the Japanese social media environment might influence the way sentiments are expressed and received in relation to {search_query}.

Provide a nuanced perspective on the overall community sentiment, highlighting the majority opinion while also considering dissenting views. The essay should be a well-structured, cohesive piece that flows logically from one point to the next.

Comments:
{highest_sentiment_comments}
{lowest_sentiment_comments}

At the end of the essay, please evaluate the overall sentiment of the comments as either positive, neutral, or negative; and assign a sentiment score from 0-10, with 0 being highly negative and 10 being highly positive. Provide your analysis in the following format:

Sentiment: <positive/neutral/negative>
Score: <0-10>"""

    # Generate response
    response = model.generate_content(system_prompt)
    return response.text

def summarize_article_with_gpt(article_text, language_selection):
    if client_gpt is None:
        print("\nOpenAIのAPIキーが見つかりません。代わりにGeminiを使用します。")
        return summarize_article_with_gemini(article_text, language_selection)

    language_options = {"1": "Japanese", "2": "English"}
    language = language_options.get(language_selection, "")
    
    try:
        response = client_gpt.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": f"You are an expert content creator with a knack for providing concise yet captivating summaries in {language}."},
                {"role": "user", "content": f"In {language}, please provide a detailed and engaging summary of the following: {article_text}. In your summary, highlight notable details, quotations, and/or statistics while aiming for a summary that's easy to read yet informative. If any additional unrelated news items, segments, warnings, alerts, stories, notifications, or advertisements arise, then ignore in your summary in favor of the primary topic: {header}. Do not reference these instructions in your response."},
            ]
        )
        return response.choices[0].message.content
    except Exception as e:
        print(f"\nOpenAI APIエラー: {str(e)}")
        print("代わりにGeminiを使用します。")
        return summarize_article_with_gemini(article_text, language_selection)

def summarize_topic_with_gpt(topcomment_text, language_selection):
    if client_gpt is None:
        print("\nOpenAIのAPIキーが見つかりません。代わりにGeminiを使用します。")
        return summarize_topic_with_gemini(topcomment_text, language_selection)

    language_options = {"1": "Japanese", "2": "English"}
    language = language_options.get(language_selection, "")
    
    try:
        response = client_gpt.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": f"You are a social media analyst who excels at explaining the meaning of posts in {language}."},
                {"role": "user", "content": f"In {language}, please explain the overall meaning and likely intent of the post, combining insights from its title, {header}, and the accompanying comment, {topcomment_text}. Avoid restating the title and comment or using introductory phrases."},
            ]
        )
        return response.choices[0].message.content
    except Exception as e:
        print(f"\nOpenAI APIエラー: {str(e)}")
        print("代わりにGeminiを使用します。")
        return summarize_topic_with_gemini(topcomment_text, language_selection)

def evaluate_sentiment_with_gpt(search_query, highest_sentiment_comments, lowest_sentiment_comments, language_selection, comments_to_analyze):
    if client_gpt is None:
        print("\nOpenAIのAPIキーが見つかりません。代わりにGeminiを使用します。")
        return evaluate_sentiment_with_gemini(search_query, highest_sentiment_comments, lowest_sentiment_comments, language_selection, comments_to_analyze)

    language_options = {"1": "Japanese", "2": "English"}
    language = language_options.get(language_selection, "")
    
    try:
        response = client_gpt.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": f"You are a helpful Japanese assistant trained to analyze comments from the online forum GirlsChannel.net related to {search_query} and evaluate the sentiments of the users. Provide a nuanced, integrated essay of approximately 500 words summarizing the prevailing sentiments based on the comments provided, highlighting themes, emotions, contrasts, and sociocultural context without using numbered sections."},
                {"role": "user", "content": f"In {language}, provide a comprehensive analysis and summary of the overall sentiment on {search_query} based on the following dataset aggregated from GirlsChannel.net. This dataset includes the {comments_to_analyze} most upvoted comments {highest_sentiment_comments} alongside the {comments_to_analyze} most downvoted comments {lowest_sentiment_comments}, reflecting the community's overall sentiment and opinion. Since the {lowest_sentiment_comments} have been voted down by users, you should interpret the prevailing opinion or sentiment to be in strong disagreement with the {lowest_sentiment_comments}. \n\nEvaluate these two sets of comments together, focusing on the predominant emotions, attitudes, topics, notable quotes, and the influence of the sociocultural context on how sentiments are expressed and received in relation to {search_query}. Provide a detailed summary encapsulating the community's stance on {search_query}. Aim to offer a nuanced perspective on the community's overall sentiment, highlighting the majority opinion while also considering dissenting views. \n\nComments:\n{highest_sentiment_comments}\n{lowest_sentiment_comments}\n. At the end of the essay, please evaluate the overall sentiment of the comments as either positive, neutral, or negative; and assign a sentiment score from 0-10, with 0 being highly negative and 10 being highly positive. Provide your analysis in the following format:\n\nSentiment: <positive/neutral/negative>\nScore: <0-10>"},
            ]
        )
        return response.choices[0].message.content
    except Exception as e:
        print(f"\nOpenAI APIエラー: {str(e)}")
        print("代わりにGeminiを使用します。")
        return evaluate_sentiment_with_gemini(search_query, highest_sentiment_comments, lowest_sentiment_comments, language_selection, comments_to_analyze)

def summarize_topic_with_anthropic(topcomment_text, language_selection):
    if client_anthropic is None:
        print("\nAnthropicのAPIキーが見つかりません。代わりにGeminiを使用します。")
        return summarize_topic_with_gemini(topcomment_text, language_selection)
        
    language_options = {"1": "Japanese", "2": "English"}
    language = language_options.get(language_selection, "")
    system_prompt = f"You are a social media analyst who excels at explaining the meaning of posts in {language}."
    
    try:
        message = client_anthropic.messages.create(
            model="claude-3-7-sonnet-latest",
            max_tokens=2048,
            temperature=0.2,
            system=system_prompt,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "text",
                            "text": f"In {language}, please explain the overall meaning and likely intent of the post, combining insights from its title, {header}, and the accompanying comment, {topcomment_text}. Avoid restating the title and comment or using introductory phrases."
                        }
                    ]
                }
            ]
        )
        return message.content[0].text
    except Exception as e:
        print(f"\nAnthropic APIエラー: {str(e)}")
        print("代わりにGeminiを使用します。")
        return summarize_topic_with_gemini(topcomment_text, language_selection)

def summarize_article_with_anthropic(article_text, language_selection):
    if client_anthropic is None:
        print("\nAnthropicのAPIキーが見つかりません。代わりにGeminiを使用します。")
        return summarize_article_with_gemini(article_text, language_selection)
        
    language_options = {"1": "Japanese", "2": "English"}
    language = language_options.get(language_selection, "")
    system_prompt = f"You are an expert content creator with a knack for providing concise yet captivating summaries in {language}."
    
    try:
        message = client_anthropic.messages.create(
            model="claude-3-7-sonnet-latest",
            max_tokens=2048,
            temperature=0.2,
            system=system_prompt,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "text",
                            "text": f"In {language}, please provide a detailed and engaging summary of the following: {article_text}. In your summary, highlight notable details, quotations, and/or statistics while aiming for a summary that's easy to read yet informative. If any additional unrelated news items, segments, warnings, alerts, stories, notifications, or advertisements arise, then ignore in your summary in favor of the primary topic: {header}. Do not reference these instructions in your response."
                        }
                    ]
                }
            ]
        )
        return message.content[0].text
    except Exception as e:
        print(f"\nAnthropic APIエラー: {str(e)}")
        print("代わりにGeminiを使用します。")
        return summarize_article_with_gemini(article_text, language_selection)

def evaluate_sentiment_with_anthropic(search_query, highest_sentiment_comments, lowest_sentiment_comments, language_selection, comments_to_analyze):
    if client_anthropic is None:
        print("\nAnthropicのAPIキーが見つかりません。代わりにGeminiを使用します。")
        return evaluate_sentiment_with_gemini(search_query, highest_sentiment_comments, lowest_sentiment_comments, language_selection, comments_to_analyze)
        
    language_options = {"1": "Japanese", "2": "English"}
    language = language_options.get(language_selection, "")
    system_prompt = f"You are a helpful Japanese assistant trained to analyze comments from the online forum GirlsChannel.net related to {search_query}. Your task is to write a comprehensive, cohesive essay in {language} that evaluates the sentiments expressed by the users. The essay should integrate the main themes, emotions, and attitudes present in the comments, providing a nuanced perspective on the overall community sentiment. Aim to write a well-structured essay of approximately 500 words, focusing on the predominant opinions while also considering dissenting views. The essay should flow smoothly, without being divided into distinct sections."
    
    try:
        message = client_anthropic.messages.create(
            model="claude-3-7-sonnet-latest",
            max_tokens=4096,
            temperature=0.3,
            system=system_prompt,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "text",
                            "text": f"In {language}, please write a comprehensive, cohesive essay of approximately 500 words analyzing and summarizing the overall sentiment on {search_query} based on the following dataset aggregated from GirlsChannel.net. The dataset includes the {comments_to_analyze} most upvoted comments {highest_sentiment_comments} alongside the {comments_to_analyze} most downvoted comments {lowest_sentiment_comments}, reflecting the community's overall sentiment and opinion. Since the {lowest_sentiment_comments} have been voted down by users, interpret the prevailing opinion or sentiment to be in strong disagreement with those comments.\n\nYour essay should flow smoothly from one idea to the next, without distinct sections or headers. Evaluate the upvoted and downvoted comments together, integrating the sentiments and themes they express to discern the collective opinion on {search_query}. Focus on the predominant emotions, attitudes, and topics to understand what aspects of {search_query} resonate most with the GirlsChannel.net online community.\n\nAddress the main themes or topics that emerge from the comments, the prevailing emotions and attitudes expressed by the community, how the most upvoted and downvoted comments differ in their sentiments, any notable quotes or specific comments that encapsulate the overall sentiment, and how the sociocultural context of the Japanese social media environment might influence the way sentiments are expressed and received in relation to {search_query}.\n\nProvide a nuanced perspective on the overall community sentiment, highlighting the majority opinion while also considering dissenting views. The essay should be a well-structured, cohesive piece that flows logically from one point to the next.\n\nComments:\n{highest_sentiment_comments}\n{lowest_sentiment_comments}\n. At the end of the essay, please evaluate the overall sentiment of the comments as either positive, neutral, or negative; and assign a sentiment score from 0-10, with 0 being highly negative and 10 being highly positive. Provide your analysis in the following format:\n\nSentiment: <positive/neutral/negative>\nScore: <0-10>"
                        }
                    ]
                }
            ]
        )
        return message.content[0].text
    except Exception as e:
        print(f"\nAnthropic APIエラー: {str(e)}")
        print("代わりにGeminiを使用します。")
        return evaluate_sentiment_with_gemini(search_query, highest_sentiment_comments, lowest_sentiment_comments, language_selection, comments_to_analyze)

# Force cleanup of gRPC resources
def cleanup():
    try:
        import grpc
        grpc.shutdown()
    except:
        pass

# Function to sanitize search query term or URL
def sanitize_filename(filename, menu_choice):
    if menu_choice == '1':
        # Parse the URL to get the text between "/" from the end of the string
        parts = filename.split('/')
        if len(parts) > 1:
            filename = parts[-2]
        else:
            filename = parts[-1]
    elif menu_choice == '2':
        # Leave the search query unchanged
        pass
    else:
        raise ValueError("Invalid menu choice")

    # Remove invalid characters from the filename
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        filename = filename.replace(char, '_')
    
    return filename

# Function to construct the search URL based on user input
def construct_search_url(search_query, date_selection):
    base_url = "https://girlschannel.net/topics/search/?q={}&date={}"
    date_options = {"1": "", "2": "y", "3": "m", "4": "w"}
    date_param = date_options.get(date_selection, "")
    return base_url.format(search_query, date_param)

# Function to make a request to the URL and parse the HTML for URLs and comment count
def fetch_topic_urls(search_url):
    response = requests.get(search_url, headers = headers)
    if response.status_code == 200:
        soup = BeautifulSoup(response.text, 'html.parser')
        topic_list = soup.find('ul', class_='topic-list')
        if topic_list:
            links = topic_list.find_all('a', href=True)
            base_url = "https://girlschannel.net"
            topic_urls = []
            comment_totals = []
            for link in links:
                url = urljoin(base_url, link['href'])
                topic_urls.append(url)
                comment_span = link.find('span', class_='icon-comment')
                if comment_span:
                    comment_total = comment_span.find_next_sibling('span').text
                    comment_total = comment_total.replace('コメント', '').strip()
                    comment_total = int(comment_total)
                else:
                    comment_total = 0
                comment_totals.append(comment_total)
            return topic_urls, comment_totals
    return []

def calculate_comments_to_analyze(total_comments_in_group):
    # Constants from the linear relationship
    slope = 0.0091
    intercept = 9.09
    # Calculate the number of comments to analyze
    comments_to_analyze = slope * total_comments_in_group + intercept
    # Round to the nearest whole number
    return round(comments_to_analyze)

def load_data(data):
    try:
        data = pd.DataFrame(columns=['Plus', 'Minus'])
        
        # Convert 'Plus' and 'Minus' columns to numeric
        data['Plus'] = pd.to_numeric(plus_minus_df['Plus'], errors='coerce')
        data['Minus'] = pd.to_numeric(plus_minus_df['Minus'], errors='coerce')
        
        data['Minus'] = data['Minus'].abs()  # Convert 'Minus' from negative to positive values
        data['Upvote Ratio'] = data['Plus'] / (data['Plus'] + data['Minus'])  # Calculate the upvote ratio
        return data
    except Exception as e:
        # If there's an error during the load, print the error
        print(f"An error occurred: {e}")
        return None
    
def get_top_bottom_data(data, comments_to_analyze):
    highest_upvoted = data.nlargest(comments_to_analyze, 'Plus')
    # print(highest_upvoted)
    lowest_downvoted = data.nlargest(comments_to_analyze, 'Minus')
    # print(lowest_downvoted)
    return highest_upvoted, lowest_downvoted

def calculate_consistency(data):
    # Calculate mean, median, and standard deviation
    mean_ratio = data['Upvote Ratio'].mean()
    median_ratio = data['Upvote Ratio'].median()
    std_dev = data['Upvote Ratio'].std()
    
    # Normalize mean and median closeness (scale from 0 to 1, where 1 is best)
    mean_median_closeness = 1 - abs(mean_ratio - median_ratio) / max(mean_ratio, median_ratio)
    
    # Normalize standard deviation (lower is better)
    std_dev_score = 1 - min(std_dev, 1)
    
    # Average the scores for overall consistency score
    consistency_score = (mean_median_closeness + std_dev_score) / 2
    
    # Scale the score to 0-10
    consistency_score *= 10
    
    return consistency_score, mean_ratio, median_ratio, std_dev

def calculate_overall_score(highest_score, lowest_score):
    # Calculate the overall consistency score as the average of the two scores
    overall_score = (highest_score + lowest_score) / 2
    return overall_score

def create_custom_palette(num_colors):
    # Create a custom color palette from dark blue to a medium blue
    colors = ["#08306b", "#08519c", "#2171b5", "#4292c6", "#6baed6", "#8bb4d9", "#a6c8e8", "#c4def7"]
    cmap = LinearSegmentedColormap.from_list("custom_blue", colors, N=num_colors)
    return [cmap(i) for i in range(cmap.N)]

def plot_results(header, url, comments_to_analyze, menu_choice, sanitized_query, highest_upvoted, lowest_downvoted, highest_score, lowest_score, overall_score, output_folder, i=None):
    # Create figure with Japanese fonts
    fig, font_prop = create_figure_with_japanese_fonts()
    
    # Create subplots
    sns.set_theme(style="whitegrid")
    fig, axs = plt.subplots(2, 2, figsize=(14, 10))
    
    # Main title and URL
    fig.suptitle(f"{header[:40]}", fontsize=16, fontweight='bold', color='black', fontproperties=font_prop)
    plt.figtext(0.5, 0.945, url, ha='center', va='center', fontsize=12)

    # Plot highest upvoted comments
    num_highest_points = len(highest_upvoted)
    highest_intervals = max(1, num_highest_points // 10)
    highest_xticks = range(1, num_highest_points + 1, highest_intervals)
    highest_xticklabels = [str(x) for x in highest_xticks]

    sns.barplot(ax=axs[0, 0], x=range(1, num_highest_points + 1), y=highest_upvoted['Plus'], 
                palette=create_custom_palette(num_highest_points))
    axs[0, 0].set_xticks(highest_xticks)
    axs[0, 0].set_xticklabels(highest_xticklabels)
    axs[0, 0].set_title('最も高評価のコメント', fontproperties=font_prop)
    axs[0, 0].set_xlabel(f'最も高評価の{comments_to_analyze}件のコメント', fontproperties=font_prop)
    axs[0, 0].set_ylabel('高評価', fontproperties=font_prop)

    # Plot lowest downvoted comments
    num_lowest_points = len(lowest_downvoted)
    lowest_intervals = max(1, num_lowest_points // 10)
    lowest_xticks = range(1, num_lowest_points + 1, lowest_intervals)
    lowest_xticklabels = [str(x) for x in lowest_xticks]

    sns.barplot(ax=axs[0, 1], x=range(1, num_lowest_points + 1), y=lowest_downvoted['Minus'], 
                palette=create_custom_palette(num_lowest_points))
    axs[0, 1].set_xticks(lowest_xticks)
    axs[0, 1].set_xticklabels(lowest_xticklabels)
    axs[0, 1].set_title('最も低評価のコメント', fontproperties=font_prop)
    axs[0, 1].set_xlabel(f'最も低評価の{comments_to_analyze}件のコメント', fontproperties=font_prop)
    axs[0, 1].set_ylabel('低評価', fontproperties=font_prop)

    # Plot consistency scores
    sns.barplot(ax=axs[1, 0], x=['高評価', '低評価'], y=[highest_score, lowest_score], 
                palette=['#FD7588', '#8CBBD9'])
    axs[1, 0].set_title('一貫性', fontproperties=font_prop)
    axs[1, 0].set_ylabel('一貫性の度合（0-10）', fontproperties=font_prop)
    
    # Set ticks explicitly before setting labels
    axs[1, 0].set_xticks([0, 1])  # Set tick positions
    axs[1, 0].set_xticklabels(['高評価', '低評価'], fontproperties=font_prop)

    # Add value labels on bars
    for bar, score in zip(axs[1, 0].patches, [highest_score, lowest_score]):
        height = bar.get_height()
        axs[1, 0].text(bar.get_x() + bar.get_width()/2., height,
                      f'{score:.2f}', ha='center', va='bottom', fontproperties=font_prop)

    # Plot overall score as pie chart
    wedges, texts, autotexts = axs[1, 1].pie([overall_score, 10 - overall_score],
                                            labels=['合意の度合', '不一致の度合'],
                                            autopct='%1.1f%%',
                                            colors=['#FD7588', '#8CBBD9'])
    
    # Apply font properties to pie chart labels and percentages
    plt.setp(texts, fontproperties=font_prop)
    plt.setp(autotexts, fontproperties=font_prop)
    axs[1, 1].set_title('全体的な合意度', fontproperties=font_prop)

    plt.tight_layout()

    # Save plot
    if menu_choice == '1':
        output_plot = os.path.join(output_folder, f"plot_{sanitized_query}.png")
    elif menu_choice == '2':
        output_plot = os.path.join(output_folder, f"plot_{sanitized_query}.{TopicCounter.plot_results_count}.png")
        TopicCounter.plot_results_count += 1
    else:
        sys.exit(1)

    plt.savefig(output_plot, dpi=300, bbox_inches='tight')
    plt.show()

def plot_comment_frequency(df, output_folder, sanitized_query, dates):
    # Create figure with Japanese fonts
    fig, font_prop = create_figure_with_japanese_fonts()
    plt.figure(figsize=(12, 6))
    
    # Prepare data
    comment_counts = df.groupby(pd.Grouper(key='Date', freq='h')).size().reset_index(name='Comment Count')

    if len(comment_counts) > 1000:
        nth_point = len(comment_counts) // 1000
        comment_counts = comment_counts.iloc[::nth_point]

    start_time = dates['earliest']
    end_time = dates['latest']
    
    # Apply 30-day limit due to platform restriction
    max_duration = timedelta(days=30)
    if (end_time - start_time) > max_duration:
        end_time = start_time + max_duration
        # Filter comment_counts to only include data within the 30-day window
        comment_counts = comment_counts[comment_counts['Date'] <= end_time]

    if (end_time - start_time) < pd.Timedelta(hours=6):
        start_time -= timedelta(hours=1)
        end_time += timedelta(hours=1)

    time_range = end_time - start_time

    # Determine time format and locator
    if time_range <= timedelta(days=1):
        time_format = '%H時'
        locator = mdates.HourLocator(interval=2)
    elif time_range <= timedelta(days=7):
        time_format = '%d日'
        locator = mdates.DayLocator(interval=1)
    elif time_range <= timedelta(days=30):
        time_format = '%d日'
        locator = mdates.DayLocator(interval=2)
    elif time_range <= timedelta(days=365):
        time_format = '%m月'
        locator = mdates.MonthLocator(interval=1)
    else:
        time_format = '%Y年'
        locator = mdates.YearLocator(interval=1)

    # Create plot
    ax = sns.lineplot(x='Date', y='Comment Count', data=comment_counts, color='#FD7588', 
                     linewidth=2, alpha=0.7)
    
    plt.xlabel(f"{start_time.strftime('%Y年%m月%d日 %H:%M')} - {end_time.strftime('%Y年%m月%d日 %H:%M')}", 
              fontproperties=font_prop, fontsize=10, labelpad=20)
    plt.title('コメントの頻度', fontproperties=font_prop, fontsize=14)
    plt.ylabel('コメント数', fontproperties=font_prop, fontsize=12)

    # Set up time axis
    formatter = mdates.DateFormatter(time_format)
    ax.xaxis.set_major_locator(locator)
    ax.xaxis.set_major_formatter(formatter)
    ax.set_xlim(start_time, end_time)

    # Apply font properties to tick labels
    for label in ax.get_xticklabels():
        label.set_fontproperties(font_prop)
    for label in ax.get_yticklabels():
        label.set_fontproperties(font_prop)

    plt.xticks(rotation=0, fontsize=10)
    plt.grid(True)
    plt.tight_layout()

    # Save plot
    if menu_choice == '1':
        output_plot_lineplot = os.path.join(output_folder, f"lineplot_{sanitized_query}.png")
    elif menu_choice == '2':
        output_plot_lineplot = os.path.join(output_folder, f"lineplot_{sanitized_query}.{TopicCounter.plot_comment_frequency_count}.png")
        TopicCounter.plot_comment_frequency_count += 1
    else:
        sys.exit(2)

    plt.savefig(output_plot_lineplot, dpi=300, bbox_inches='tight')
    plt.show()

async def process_url(url, worksheet, comment_total):
    async with aiohttp.ClientSession() as session:
        while url:
            new_url, header, comment_number = await fetch_and_process_page(url, worksheet, comment_total, session)
            if not new_url or new_url == url:  # Prevent infinite loop
                break
            url = new_url
            await asyncio.sleep(1)  # Add a delay between requests
    return header, comment_number

async def fetch_and_process_page(url, worksheet, comment_total, session=None):
    """
    Asynchronously fetch and process a webpage, extracting comments and related data.
    """
    # Add total_processed as a parameter to track progress across pages
    global total_processed
    if not hasattr(fetch_and_process_page, 'total_processed'):
        fetch_and_process_page.total_processed = 0

    try:
        if session is None:
            async with aiohttp.ClientSession() as session:
                return await _process_page(session, url, worksheet, comment_total, headers)
        else:
            return await _process_page(session, url, worksheet, comment_total, headers)
    except Exception as e:
        print(f"\nError in fetch_and_process_page: {str(e)}")
        return None, None, None

async def _process_page(session, url, worksheet, comment_total, headers):
    """Helper function to process a single page"""
    try:
        async with session.get(url, headers=headers) as response:
            if response.status != 200:
                print(f"\nError: HTTP {response.status} when accessing {url}")
                return None, None, None

            content = await response.text()
            soup = BeautifulSoup(content, 'html.parser')
            
            header_div = soup.find('div', class_='head-area')
            if not header_div or not header_div.find('h1'):
                print("\nError: Could not find article header")
                return None, None, None
                
            header = header_div.find('h1').text.strip()
            
            # Only print title and total once, and store it in a static variable
            if not hasattr(_process_page, 'title_printed'):
                _process_page.title_printed = True
                print(f"\n記事タイトル: {header} (コメント数: {comment_total})")

            comments = soup.find_all('li', class_='comment-item')
            if not comments:
                print("\nError: No comments found")
                return None, None, None

            comment_number = None
            processed_count = 0
            
            for comment in comments:
                try:
                    comment_number = comment.find('p', class_='info').get_text(strip=True).split('.')[0]
                    processed_count += 1
                    
                    # Process comment data
                    date = comment.find('a', rel='nofollow').get_text(strip=True)
                    date = re.sub(r'\(.*?\)', '', date)
                    date = datetime.datetime.strptime(date, '%Y/%m/%d %H:%M:%S')
                    
                    wrap_plus = comment.find('div', class_='icon-rate-wrap-plus').find('p').get_text(strip=True)
                    wrap_minus = comment.find('div', class_='icon-rate-wrap-minus').find('p').get_text(strip=True)
                    wrap_total = int(wrap_plus) + int(wrap_minus)

                    # Process comment body
                    body_div = comment.find('div', class_='body')
                    for span in body_div.find_all('span', class_='res-anchor'):
                        span.decompose()
                    comment_body = body_div.get_text(strip=True, separator=' ').replace('\n', ' ')

                    if not comment_body.strip() or comment_body.startswith("出典"):
                        img_tag = body_div.select_one('div > img')
                        if img_tag and 'data-src' in img_tag.attrs:
                            comment_body = str(img_tag['data-src'])
                        else:
                            comment_body = '「画像のURLが見分けられない。」'

                    # Process sub-comments
                    sub_comments = ""
                    comment_url_title_div = comment.find('div', class_='comment-url-title')
                    if comment_url_title_div:
                        urls = set([a['href'] for a in comment_url_title_div.find_all('a', href=True)])
                        if urls:
                            sub_comments = '\n'.join(urls)

                    # Clean comment number 1
                    if comment_number == '1':
                        comment_body = re.sub(r'出典：.*?(?=\s|$)', '', comment_body)

                    comment_number_formatted = comment_number.zfill(4)

                    # Write to worksheet
                    if comment_number == '1':
                        worksheet.append([comment_number_formatted, date, wrap_plus, wrap_minus, wrap_total, comment_body, sub_comments, header])
                    else:
                        worksheet.append([comment_number_formatted, date, wrap_plus, wrap_minus, wrap_total, comment_body, sub_comments, ''])
                    
                    # Update progress without reprinting the title
                    processed_count += 1
                    total_processed = int(comment_number) if comment_number else processed_count
                    progress = (total_processed / comment_total) * 100
                    print(f"\rコメント処理中: {total_processed}/{comment_total} ({progress:.1f}%)", end='')
                    
                except Exception as e:
                    print(f"\nError processing comment {processed_count}: {str(e)}")
                    continue

            # Look for next page without printing message
            next_page_url = None
            pager_area = soup.find('ul', class_='pager-topic')
            if pager_area:
                next_page_link = pager_area.find('li', class_='next')
                if next_page_link and next_page_link.find('a'):
                    next_page_url = urljoin(url, next_page_link.find('a')['href'])

            return next_page_url, header, comment_number

    except Exception as e:
        print(f"\nError processing page {url}: {str(e)}")
        return None, None, None

async def process_additional_comments(res_count_url, parent_comment_number, worksheet, session):
    """
    Process additional comments asynchronously
    """
    base_url = "https://girlschannel.net"
    res_count_full_url = urljoin(base_url, res_count_url)
    
    try:
        async with session.get(res_count_full_url) as response:
            content = await response.text()
            res_count_soup = BeautifulSoup(content, 'html.parser')
            res_count_comments = res_count_soup.find_all('li', class_='comment-item')

            if len(res_count_comments) > 1:
                res_count_comments = res_count_comments[1:]
                subcomment_counter = 1

                for res_count_comment in res_count_comments:
                    # Extract and process comment data
                    res_count_date = res_count_comment.find('a', rel='nofollow').get_text(strip=True)
                    res_count_date = re.sub(r'\(.*?\)', '', res_count_date)
                    res_count_date = datetime.datetime.strptime(res_count_date, '%Y/%m/%d %H:%M:%S')
                    
                    res_count_wrap_plus = res_count_comment.find('div', class_='icon-rate-wrap-plus').find('p').get_text(strip=True)
                    res_count_wrap_minus = res_count_comment.find('div', class_='icon-rate-wrap-minus').find('p').get_text(strip=True)
                    res_count_sentiment = int(res_count_wrap_plus) + int(res_count_wrap_minus)

                    body_div = res_count_comment.find('div', class_='body')
                    for span in body_div.find_all('span', class_='res-anchor'):
                        span.decompose()
                    res_count_comment_body = body_div.get_text(strip=True)

                    if not res_count_comment_body.strip() or res_count_comment_body.startswith("出典"):
                        img_tag = body_div.select_one('div > img')
                        if img_tag and 'data-src' in img_tag.attrs:
                            res_count_comment_body = str(img_tag['data-src'])
                        else:
                            res_count_comment_body = '「画像のURLが見分けられない。」'

                    subcomment_counter_formatted = str(subcomment_counter).zfill(3)
                    worksheet.append([
                        f"{parent_comment_number}.{subcomment_counter_formatted}", 
                        res_count_date, 
                        res_count_wrap_plus, 
                        res_count_wrap_minus, 
                        res_count_sentiment, 
                        res_count_comment_body, 
                        '', 
                        ''
                    ])
                    subcomment_counter += 1

    except Exception as e:
        print(f"Error processing additional comments from {res_count_full_url}: {str(e)}")

# Update the run_async function to use nest_asyncio
import nest_asyncio
nest_asyncio.apply()

def run_async(coroutine):
    """Helper function to run async code in any context."""
    try:
        loop = asyncio.get_event_loop()
        return loop.run_until_complete(coroutine)
    except Exception as e:
        print(f"Error in run_async: {str(e)}")
        return None, None

# Update the process_url function to be less verbose
async def process_url(url, worksheet, comment_total):
    """Process URL and return header and comment number"""
    try:
        timeout = aiohttp.ClientTimeout(total=30)
        async with aiohttp.ClientSession(timeout=timeout) as session:
            current_url = url
            header = None
            comment_number = None
            
            while current_url:
                new_url, new_header, new_comment_number = await fetch_and_process_page(current_url, worksheet, comment_total, session)
                
                if new_header and header is None:
                    header = new_header
                if new_comment_number:
                    comment_number = new_comment_number
                
                if not new_url or new_url == current_url:
                    break
                    
                current_url = new_url
                await asyncio.sleep(1)
            
            return header, comment_number
    except Exception as e:
        print(f"\nError in process_url: {str(e)}")
        return None, None

# Update the main function to be less verbose
async def main(url, worksheet, comment_total):
    """Main async function to handle URL processing"""
    try:
        print("\nコメントの取得を開始します...")
        result = await process_url(url, worksheet, comment_total)
        print("\n完了しました")
        return result
    except Exception as e:
        print(f"\nError in main: {str(e)}")
        return None, None

if __name__ == "__main__":
    menu_choice, *args = get_user_input()
    
    # Create a new Excel workbook
    workbook = Workbook()
    
    if menu_choice == '1':
        url, language_selection, comment_total, summary_ai, sentiment_ai = args
        sanitized_query = sanitize_filename(url, menu_choice)
        output_file_name = os.path.join(output_folder, f"output_{sanitized_query}.txt")
        output_file = open(output_file_name, "w", encoding="utf-8")
        excel_file_name = os.path.join(output_folder, f"comments_{sanitized_query}.xlsx")
        num_topics_found = 1
        
        print("\n*** GirlsinSight ***", file=output_file)
        print(f"\nURL: {url}", file=output_file)
        print(f"\n入力されたURL: {url}")
        
        # Create a new worksheet for the current URL
        worksheet = workbook.create_sheet(title="Topic 1")
        
        # Write the header row to the worksheet
        worksheet.append(['Comment Number', 'Date', 'Plus', 'Minus', 'Total', 'Comment', 'URLs', 'Header'])
        
        print("\nURLの処理を開始します...")
        # Process the URL asynchronously
        header, comment_number = run_async(main(url, worksheet, comment_total))
        if header:
            search_query = header
            print(f"\n処理が完了しました: {header}")
        else:
            print("\nURLの処理に失敗しました")
            sys.exit(1)
    
    elif menu_choice == '2':
        search_query, topic_urls, comment_totals, num_topics_found, language_selection, summary_ai, sentiment_ai = args
        sanitized_query = sanitize_filename(search_query, menu_choice)
        output_file_name = os.path.join(output_folder, f"output_{sanitized_query}.txt")
        output_file = open(output_file_name, "w", encoding="utf-8")
        excel_file_name = os.path.join(output_folder, f"comments_{sanitized_query}.xlsx")

        count = 1
        
        print("\n*** GirlsInsight ***", file=output_file)
        print(f"検索キーワード: {search_query}", file=output_file)
        print(f"検索キーワード: {search_query}")
        print(f"トピックの総数: {num_topics_found}", file=output_file)
        
        for url, comment_total in zip(topic_urls, comment_totals):
            print(f"{count}. {url}", file=output_file)
            print(url)

            # Create a new worksheet for the current URL
            worksheet = workbook.create_sheet(title=f"Topic {count}")

            # Write the header row to the worksheet
            worksheet.append(['Comment Number', 'Date', 'Plus', 'Minus', 'Total', 'Comment', 'URLs', 'Header'])

            print(f"\nURLの処理を開始します: {url}")
            # Process the URL asynchronously
            header, comment_number = run_async(main(url, worksheet, comment_total))
            if header:
                print(f"\n処理が完了しました: {header}")
            else:
                print(f"\nURLの処理に失敗しました: {url}")
                continue

            print(f"\n{count}件のトピックの読込みに成功しました。\n")
            count += 1
            print(file=output_file)
    
    # Remove the default sheet
    default_sheet = workbook['Sheet']
    workbook.remove(default_sheet)
    
    # Save and read the Excel file into a DataFrame
    workbook.save(excel_file_name)
    workbook.close()

    df = pd.read_excel(excel_file_name, sheet_name=None)

    # Combine all worksheets into a single DataFrame
    df = pd.concat(df.values(), ignore_index=True)
    
    # Verify the necessary columns exist in the DataFrame
    required_columns = ['Comment Number', 'Date', 'Plus', 'Minus', 'Total', 'Comment', 'URLs', 'Header']
    if not all(column in df.columns for column in required_columns):
        raise ValueError(f"Excel file must contain the columns: {', '.join(required_columns)}")

    # Create a new DataFrame to store the comment groups
    comment_groups_df = pd.DataFrame(columns=['Date', 'Title', 'Topic', 'Highest Sentiment Comments', 'Lowest Sentiment Comments'])

    # Create a new DataFrame to store the total number of comments in each comment group
    comment_group_totals_df = pd.DataFrame(columns=['Total Comments'])
    
    #Create a new DataFrame to store the Plus and Minus
    plus_minus_df = pd.DataFrame(columns=['Plus', 'Minus'])
    
    # Create an Excel workbook
    workbook = Workbook()
        
    # Create numbered comment groups
    start_index = 0
    for i in range(num_topics_found):
        if start_index < len(df):
            # Find the next occurrence of comment number 1 with a header
            next_comment_number_1 = df.loc[(df['Comment Number'] == 1) & (df.index > start_index) & (df['Header'].notna())]
    
            if not next_comment_number_1.empty:
                end_index = next_comment_number_1.index[0] - 1
            else:
                end_index = len(df) - 1  # Set end_index to the last index of the DataFrame
    
            group_df = df.iloc[start_index:end_index+1]
            dates = process_dates(group_df)

            group_comments = ''
            group_date = group_df.iloc[0]['Date']
            group_title = group_df.iloc[0]['Comment']
            header = group_df.iloc[0]['Header']
            group_plus = group_df.iloc[0]['Plus']
            group_minus = group_df.iloc[0]['Minus']
    
            total_comments = len(group_df)

            for _, row in group_df.iterrows():
                if pd.notna(row['Comment']):
                    group_comments += f"{row['Comment']}\n"
    
            comments_to_analyze = calculate_comments_to_analyze(total_comments)
    
            group_plus_minus = group_df[['Plus', 'Minus']]
            plus_minus_df = pd.concat([plus_minus_df, group_plus_minus], ignore_index=True)
    
            data = load_data(group_plus_minus)
            highest_upvoted, lowest_downvoted = get_top_bottom_data(data, comments_to_analyze)
                        
            highest_sentiment_comments = group_df.nlargest(comments_to_analyze, 'Total')['Comment'].tolist()
            lowest_sentiment_comments = group_df.nsmallest(comments_to_analyze, 'Total')['Comment'].tolist()
    
            new_row = pd.DataFrame({
                'Date': [group_date],
                'Title': [group_title],
                'Plus': [group_plus],
                'Minus': [group_minus],
                'Header': [header],
                'Topic': [group_comments],
                'Highest Sentiment Comments': [highest_sentiment_comments],
                'Lowest Sentiment Comments': [lowest_sentiment_comments]
            })
    
            new_row_totals = pd.DataFrame({
                'Total Comments': [total_comments]
            })
    
            new_row_plus_minus = pd.DataFrame({
                'Plus': [group_plus],
                'Minus': [group_minus]
            })
    
            # Exclude empty or all-NA entries from new_row before concatenation
            new_row = new_row.dropna(axis=1, how='all')
            comment_groups_df = pd.concat([comment_groups_df, new_row], ignore_index=True)
            
            # Exclude empty or all-NA entries from new_row_totals before concatenation
            new_row_totals = new_row_totals.dropna(axis=1, how='all')
            comment_group_totals_df = pd.concat([comment_group_totals_df, new_row_totals], ignore_index=True)
            
            # Exclude empty or all-NA entries from new_row_plus_minus before concatenation
            new_row_plus_minus = new_row_plus_minus.dropna(axis=1, how='all')
            plus_minus_df = pd.concat([plus_minus_df, new_row_plus_minus], ignore_index=True)
        
            # Write the group DataFrame to the worksheet
            for row in dataframe_to_rows(group_df, index=False, header=True):
                worksheet.append(row)
    
        else:
            break
    
        date = comment_groups_df.loc[i, 'Date']
        title = comment_groups_df.loc[i, 'Title']  # Retrieve the title of the comment group
        topic = comment_groups_df.loc[i, 'Topic']
        highest_sentiment_comments = comment_groups_df.loc[i, 'Highest Sentiment Comments']
        lowest_sentiment_comments = comment_groups_df.loc[i, 'Lowest Sentiment Comments']
        header = comment_groups_df.loc[i, 'Header']  # Retrieve the header for the current comment group
    
        # Define group_df for the current comment group
        group_df = df.iloc[start_index:i+1]  # Include the current row in the group
        print('')
        print(f"\nトピック {i+1}:", file=output_file)
        print(f"{header}", file=output_file)
        print('')
        print(f"トピック {i+1}:")
        print(f"{header}")
    
        # Directly fetch the first URL from the DataFrame
        url_to_process = df.loc[(df['Comment Number'] == 1) & (df.index >= start_index), 'URLs'].iloc[0]
        
        article_text = "" # Initialize article_text
        news_story_status = None # Initialize
        source_to_display_for_summary = None # Will hold the primary URL associated with the article text
        wayback_url_actually_used = None # Specifically if a wayback URL was successfully used

        # Check if the URL is not NaN
        if pd.notna(url_to_process):
            news_story_status = check_news_story_status(url_to_process)
            if news_story_status == "active":
                all_urls = create_url_list(url_to_process)
                _article_text_candidate = extract_text_from_url(all_urls, headers)
                if _article_text_candidate:
                    article_text = _article_text_candidate
                    source_to_display_for_summary = url_to_process # Active URL is the source
            elif news_story_status == "inactive":
                print("\n出典先は無効でした。")
                _found_archive_url = search_wayback_machine(url_to_process, headers)
                if _found_archive_url:
                    print("\nアーカイブ版が見つかりました。")
                    _article_text_candidate = extract_text_from_url([_found_archive_url], headers)
                    if _article_text_candidate:
                        article_text = _article_text_candidate
                        source_to_display_for_summary = url_to_process # Original URL is still the primary reference
                        wayback_url_actually_used = _found_archive_url # This archive was used
                else:
                    print("ウェイバックマシンにアーカイブ版も見つかりませんでした。")
           
            if article_text:
                summary = globals()[f"summarize_article_with_{summary_ai}"](article_text, language_selection)
                print(file=output_file)
                print()
                print("概要:", file=output_file)
                print("\n概要:")
                print(summary, file=output_file)
                print(summary)
                print(file=output_file)  # Optional: Add an empty print statement for a line break
                
                if wayback_url_actually_used:
                    print(f"Original source (inactive): {source_to_display_for_summary}", file=output_file)
                    print(f"\nOriginal source (inactive): {source_to_display_for_summary}")
                    print(f"Archived version used: {wayback_url_actually_used}", file=output_file)
                    print(f"Archived version used: {wayback_url_actually_used}")
                elif source_to_display_for_summary: # Active URL provided the text, and no wayback was involved
                    print(f"出典: {source_to_display_for_summary}", file=output_file)
                    print(f"\n出典: {source_to_display_for_summary}")
                # If both are None but article_text exists, that would be an edge case not covered by current var assignments
            else: # article_text is empty
                print("\n記事のテキストが見つかりませんでした。要約の生成を見送りました。", file=output_file)
                print("\n記事のテキストが見つかりませんでした。要約の生成を見送りました。")
                if news_story_status == "inactive":
                     print(f"Original source (inactive, no usable archive found or text extraction failed): {url_to_process}", file=output_file)
                     print(f"Original source (inactive, no usable archive found or text extraction failed): {url_to_process}")
                elif news_story_status == "active": # Active, but text extraction failed
                     print(f"Source (active, but text extraction failed): {url_to_process}", file=output_file)
                     print(f"Source (active, but text extraction failed): {url_to_process}")
        # Directly fetch the top comment from the DataFrame if url_to_process was NaN
        else:
            topcomment_text = df.loc[df['Comment Number'] == 1, 'Comment'].iloc[i]
            summary = globals()[f"summarize_topic_with_{summary_ai}"](topcomment_text, language_selection)
            print(file=output_file)
            print()
            print("概要:", file=output_file)
            print("\n概要:")
            print(summary, file=output_file)
            print(summary)
            print(file=output_file)  # Add an empty print statement for a line break
            print()  # Add an empty print statement for a line break
            print("出典がありません", file=output_file)
            print("出典がありません")
    
        next_comment_number_1 = df.loc[(df['Comment Number'] == 1) & (df.index > start_index)]
        if not next_comment_number_1.empty:
            start_index = next_comment_number_1.index[0]
        else:
            start_index = len(df)  # Set start_index to the end of the DataFrame if no more comment_number 1 occurrences
            
        print(file=output_file)  # Add an empty print statement for a line break
        print()  # Add an empty print statement for a line break    
        print(f"初コメント: {dates['earliest'].strftime('%Y年%m月%d日 %H:%M')}")
        print(f"最終コメント: {dates['latest'].strftime('%Y年%m月%d日 %H:%M')}")
        print(f"初コメント: {dates['earliest'].strftime('%Y年%m月%d日 %H:%M')}", file=output_file)
        print(f"最終コメント: {dates['latest'].strftime('%Y年%m月%d日 %H:%M')}", file=output_file)
        if dates['days'] > 0:
            print(f"期間: {dates['days']}日間{dates['hours']}時間{dates['minutes']}分{dates['seconds']}秒")
            print(f"期間: {dates['days']}日間{dates['hours']}時間{dates['minutes']}分{dates['seconds']}秒", file=output_file)
        else:
            print(f"期間: {dates['hours']}時間{dates['minutes']}分{dates['seconds']}秒")
            print(f"期間: {dates['hours']}時間{dates['minutes']}分{dates['seconds']}秒", file=output_file)
        plot_comment_frequency(df, output_folder, sanitized_query, dates)
        print(file=output_file)  # Add an empty print statement for a line break
        print()  # Add an empty print statement for a line break
        print("最も投票数の多いコメント:", file=output_file)
        print("最も投票数の多いコメント:")
    
        # Calculate the number of comments to analyze based on the total number of comments in the group
        total_comments_in_group = int(comment_group_totals_df.loc[i, 'Total Comments'])
        comments_to_analyze = calculate_comments_to_analyze(total_comments_in_group)
    
        # Get the 'Plus' and 'Minus' values for the current comment group
        group_plus_minus = group_df[['Plus', 'Minus']]
        data = load_data(group_plus_minus)
        highest_upvoted, lowest_downvoted = get_top_bottom_data(data, comments_to_analyze)
    
        # Calculate consistency scores and stats for highest and lowest
        highest_score, highest_mean, highest_median, highest_std = calculate_consistency(highest_upvoted)
        lowest_score, lowest_mean, lowest_median, lowest_std = calculate_consistency(lowest_downvoted)
    
        # Calculate the overall consistency score
        overall_score = calculate_overall_score(highest_score, lowest_score)
    
        for comment in highest_sentiment_comments[:3]:
            sentiment_score = df.loc[df['Comment'] == comment, 'Total'].values
            if len(sentiment_score) > 0:
                print(f"∙ コメント: {comment}", file=output_file)
                print(f"∙ コメント: {comment}")
                print(f"  点数: {sentiment_score[0]}", file=output_file)
                print(f"  点数: {sentiment_score[0]}")
                print(file=output_file)  # Add an empty print statement for a line break
                print()  # Add an empty print statement for a line break
            else:
                print(f"∙ コメント: {comment}", file=output_file)
                print(f"∙ コメント: {comment}")
                print(file=output_file)  # Add an empty print statement for a line break
                print()  # Add an empty print statement for a line break
        print(f"　平均比率: {highest_mean:.2f}")
        print(f"　平均比率: {highest_mean:.2f}", file=output_file)
        print(f"　中央値比率: {highest_median:.2f}")
        print(f"　中央値比率: {highest_median:.2f}", file=output_file)
        print(f"　標準偏差: {highest_std:.2f}")
        print(f"　標準偏差: {highest_std:.2f}", file=output_file)
        print(f"　合意の度合: {highest_score:.2f} (0-10)\n")
        print(f"　合意の度合: {highest_score:.2f} (0-10)\n", file=output_file)
        print("最も低評価のコメント:", file=output_file)
        print("最も低評価のコメント:")
        for comment in lowest_sentiment_comments[:3]:
            sentiment_score = df.loc[df['Comment'] == comment, 'Total'].values
            if len(sentiment_score) > 0:
                print(f"∙ コメント: {comment}", file=output_file)
                print(f"∙ コメント: {comment}")
                print(f"  点数: {sentiment_score[0]}", file=output_file)
                print(f"  点数: {sentiment_score[0]}")
                print(file=output_file)  # Add an empty print statement for a line break
                print()  # Add an empty print statement for a line break
            else:
                print(f"∙ コメント: {comment}", file=output_file)
                print(f"∙ コメント: {comment}")
                print(file=output_file)  # Add an empty print statement for a line break
                print()  # Add an empty print statement for a line break
        print(f"　平均比率: {lowest_mean:.2f}")
        print(f"　平均比率: {lowest_mean:.2f}", file=output_file)
        print(f"　中央値比率: {lowest_median:.2f}")
        print(f"　中央値比率: {lowest_median:.2f}", file=output_file)
        print(f"　標準偏差: {lowest_std:.2f}")
        print(f"　標準偏差: {lowest_std:.2f}", file=output_file)
        print(f"　不一致の度合: {lowest_score:.2f}\n")
        print(f"　不一致の度合: {lowest_score:.2f}\n", file=output_file)
        evaluation = globals()[f"evaluate_sentiment_with_{sentiment_ai}"](search_query, highest_sentiment_comments, lowest_sentiment_comments, language_selection, comments_to_analyze)
        print("評価:", file=output_file)
        print()
        print("評価:")
        print(evaluation, file=output_file)
        print(evaluation)
        print(f"\n評価されたコメント数: {comments_to_analyze*2}")
        print(f"総合コメント数: {total_comments_in_group}")
        print(f"\n評価されたコメント数: {comments_to_analyze*2}", file=output_file)
        print(f"総合コメント数: {total_comments_in_group}", file=output_file)
        print(file=output_file)  # Add an empty print statement for a line break
        print()  # Add an empty print statement for a line break
        print(f"全体的な合意度: {overall_score:.2f} (0-10)")
        print(f"全体的な合意度: {overall_score:.2f} (0-10)", file=output_file)
        print()  # Add an empty print statement for a line break
        print(file=output_file)  # Add an empty print statement for a line break

        # Plot the results
        plot_results(header, url, comments_to_analyze, menu_choice, sanitized_query, highest_upvoted, lowest_downvoted, highest_score, lowest_score, overall_score, output_folder, i=None)
        print("————————————————————————————", file=output_file)
        print("————————————————————————————")
        print(file=output_file)  # Add an empty print statement for a line break
    
        start_index = end_index + 1

# Print the date and time
print("GirlsinSight v1.0.1", file=output_file)
print("作成日時：" + formatted_datetime, file=output_file)

# Close the output file
output_file.close()
time.sleep(1)  # Add a small delay before script termination